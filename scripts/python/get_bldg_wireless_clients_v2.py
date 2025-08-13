"""Generate a wireless client report per building.

This script builds a summary of wireless client counts per building over
a configurable number of days and writes an Excel report.  It
leverages :mod:`na_utils.dnac` for API access and uses
``openpyxl`` to generate the report.  A backup copy of the report
can be stored in a separate directory.  Paths and time zones are
customisable via command line arguments.
"""

from __future__ import annotations

import argparse
import os
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, Any, List, Tuple

import pytz
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter

# Ensure project root on sys.path for na_utils
PROJECT_ROOT = Path(__file__).resolve().parents[2]
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from na_utils.dnac import get_api_response, get_auth_token


def get_site_topology() -> Dict[str, Any]:
    """Return the site topology from Catalyst Center."""
    return get_api_response("/api/v1/topology/site-topology").get("response", {})


def get_users_per_bldg(bldg_id: str, start_time_ms: int, end_time_ms: int) -> List[Dict[str, Any]]:
    """Return wireless client list for a building and time range."""
    endpoint = (
        f"/dna/data/api/v1/clients?startTime={start_time_ms}&endTime={end_time_ms}"
        f"&type=Wireless&siteHierarchyId=*{bldg_id}*"
    )
    resp = get_api_response(endpoint)
    return resp.get("response", [])


def generate_daily_time_ranges(days: int, timezone: str) -> List[Tuple[int, int, str]]:
    tz = pytz.timezone(timezone)
    ranges: List[Tuple[int, int, str]] = []
    for i in range(days):
        day = datetime.now(tz) - timedelta(days=i + 1)
        start = tz.localize(datetime(day.year, day.month, day.day, 6, 0, 0))
        end = tz.localize(datetime(day.year, day.month, day.day, 18, 0, 0))
        ranges.append((int(start.timestamp() * 1000), int(end.timestamp() * 1000), start.strftime("%Y-%m-%d")))
    return list(reversed(ranges))


def write_excel_report(summary_data: List[Dict[str, Any]], day_labels: List[str], filename: Path) -> None:
    """Write the summary report to an Excel file."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Summary"
    headers = ["Building Name"] + day_labels + ["Total Clients", "Average Clients/Day", "Peak Clients"]
    ws.append(headers)
    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 18
    for entry in summary_data:
        row = [entry['building']] + entry['daily_counts'] + [entry['total'], entry['average'], entry['peak']]
        ws.append(row)
    filename.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(filename))
    print(f"✅ Excel report written to: {filename}")


def save_report_copy(original_filepath: Path, destination_directory: Path) -> None:
    if not original_filepath.is_file():
        print(f"❌ Original file not found: {original_filepath}")
        return
    if not destination_directory.is_dir():
        destination_directory.mkdir(parents=True, exist_ok=True)
    destination_path = destination_directory / original_filepath.name
    try:
        import shutil
        shutil.copy2(original_filepath, destination_path)
        print(f"📁 Report copy saved to: {destination_path}")
    except Exception as e:
        print(f"❌ Failed to copy file: {e}")


def main() -> None:
    parser = argparse.ArgumentParser(description="Generate wireless client summary per building")
    parser.add_argument("--days", type=int, default=30, help="Number of days to report on")
    parser.add_argument("--timezone", default="US/Central", help="Timezone for reporting (e.g. US/Central)")
    parser.add_argument("--output", default="wireless_reports/building_client_summary.xlsx", help="Excel file path to write")
    parser.add_argument("--backup-dir", default=None, help="Optional directory to save a copy of the report")
    args = parser.parse_args()
    site_topology = get_site_topology()
    # Build a mapping of building ID to hierarchy name for floors
    bldg_id_map = {
        site['id']: site['groupNameHierarchy']
        for site in site_topology.get('sites', [])
        if site.get('locationType') == 'floor'
    }
    time_ranges = generate_daily_time_ranges(args.days, args.timezone)
    bldg_stats: Dict[str, List[int]] = {bldg_name: [] for bldg_name in bldg_id_map.values()}
    day_labels = [label for _, _, label in time_ranges]
    print(f"Collecting data for {args.days} days…")
    for day_index, (start_ms, end_ms, label) in enumerate(time_ranges, 1):
        print(f"Day {day_index} ({label}):")
        for bldg_id, bldg_name in bldg_id_map.items():
            try:
                clients = get_users_per_bldg(bldg_id, start_ms, end_ms)
                client_count = len(clients)
                bldg_stats[bldg_name].append(client_count)
                print(f"  {bldg_name}: {client_count} clients")
            except Exception as e:
                print(f"  Error fetching data for {bldg_name}: {e}")
                bldg_stats[bldg_name].append(0)
    summary_data: List[Dict[str, Any]] = []
    for bldg_name, daily_counts in bldg_stats.items():
        total = sum(daily_counts)
        avg = round(total / len(daily_counts), 2)
        peak = max(daily_counts)
        summary_data.append({
            "building": bldg_name,
            "daily_counts": daily_counts,
            "total": total,
            "average": avg,
            "peak": peak,
        })
        print(f"{bldg_name}: total={total}, avg/day={avg}, peak={peak}")
    report_path = Path(args.output)
    write_excel_report(summary_data, day_labels, report_path)
    if args.backup_dir:
        save_report_copy(report_path, Path(args.backup_dir))


if __name__ == "__main__":
    main()