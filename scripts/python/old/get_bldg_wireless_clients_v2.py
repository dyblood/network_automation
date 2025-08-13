import requests
from requests.auth import HTTPBasicAuth
from datetime import datetime, timedelta
import pytz
import os
import shutil
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from openpyxl.utils import get_column_letter
from env_DNAC import *

# Disable SSL warnings
requests.packages.urllib3.disable_warnings(requests.packages.urllib3.exceptions.InsecureRequestWarning)

def get_auth_token():
    url = f"{DNAC_BURL}/dna/system/api/v1/auth/token"
    response = requests.post(url, auth=(DNAC_USER, DNAC_PASS), verify=False)
    response.raise_for_status()
    return response.json()['Token']

def get_site_topology(token):
    url = f"{DNAC_BURL}/api/v1/topology/site-topology"
    headers = {'x-auth-token': token, 'Content-Type': 'application/json'}
    response = requests.get(url, headers=headers, verify=False)
    response.raise_for_status()
    return response.json()["response"]

def get_users_per_bldg(token, bldg_id, start_time_ms, end_time_ms):
    url = (
        f"{DNAC_BURL}/dna/data/api/v1/clients?"
        f"startTime={start_time_ms}&endTime={end_time_ms}&type=Wireless&siteHierarchyId=*{bldg_id}*"
    )
    headers = {'x-auth-token': token, 'Content-Type': 'application/json'}
    response = requests.get(url, headers=headers, verify=False)
    response.raise_for_status()
    return response.json()["response"]

def generate_daily_time_ranges(days, timezone='US/Central'):
    tz = pytz.timezone(timezone)
    time_ranges = []

    for i in range(days):
        day = datetime.now(tz) - timedelta(days=i + 1)
        start = tz.localize(datetime(day.year, day.month, day.day, 6, 0, 0))
        end = tz.localize(datetime(day.year, day.month, day.day, 18, 0, 0))
        time_ranges.append((int(start.timestamp() * 1000), int(end.timestamp() * 1000), start.strftime('%Y-%m-%d')))

    return list(reversed(time_ranges))

def write_excel_report(summary_data, days, base_path='C:\\Users\\devon.d.youngblood\\OneDrive - US Army\\Desktop\\youngblood_netops\\wireless_report\\'):
    # Extract date range
    start_date_str = days[0].replace('-', '')
    end_date_str = days[-1].replace('-', '')

    # Construct filename with date range
    filename = f"{base_path}building_client_summary ({start_date_str} - {end_date_str}).xlsx"

    wb = Workbook()
    ws = wb.active
    ws.title = "Client Summary"

    # Headers
    headers = ["Building Name"] + days + ["Total Clients", "Average Clients/Day", "Peak Clients"]
    ws.append(headers)

    header_font = Font(bold=True)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = header_font
        ws.cell(row=1, column=col).alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(col)].width = 18

    # Data Rows
    for entry in summary_data:
        row = [entry['building']] + entry['daily_counts'] + [entry['total'], entry['average'], entry['peak']]
        ws.append(row)

    wb.save(filename)
    print(f"\n✅ Excel report written to: {filename}")
    return filename

def save_report_copy(original_filepath, destination_directory):
    """
    Saves a copy of the Excel report to another directory.

    Parameters:
    - original_filepath (str): Full path to the original Excel file
    - destination_directory (str): Directory to copy the file to
    """
    if not os.path.isfile(original_filepath):
        print(f"❌ Original file not found: {original_filepath}")
        return

    if not os.path.isdir(destination_directory):
        print(f"❌ Destination directory does not exist: {destination_directory}")
        return

    filename = os.path.basename(original_filepath)
    destination_path = os.path.join(destination_directory, filename)

    try:
        shutil.copy2(original_filepath, destination_path)
        print(f"📁 Report copy saved to: {destination_path}")
    except Exception as e:
        print(f"❌ Failed to copy file: {e}")


def main():
    token = get_auth_token()
    site_topology = get_site_topology(token)

    bldg_id_map = {
        site['id']: site['groupNameHierarchy']
        for site in site_topology.get('sites', [])
        if site.get('locationType') == 'floor'
    }

    time_ranges = generate_daily_time_ranges(29)
    bldg_stats = {bldg_name: [] for bldg_name in bldg_id_map.values()}
    day_labels = [label for _, _, label in time_ranges]

    print("\nCollecting data for each building over 30 days...\n")

    for day_index, (start_ms, end_ms, label) in enumerate(time_ranges, 1):
        print(f"Day {day_index} ({label}):")
        for bldg_id, bldg_name in bldg_id_map.items():
            try:
                clients = get_users_per_bldg(token, bldg_id, start_ms, end_ms)
                client_count = len(clients)
                bldg_stats[bldg_name].append(client_count)
                print(f"  {bldg_name}: {client_count} clients")
            except Exception as e:
                print(f"  Error fetching data for {bldg_name}: {e}")
                bldg_stats[bldg_name].append(0)

    print("\nSummary Report (Last 30 Days 0600–1800 CST):")
    print("-" * 60)

    summary_data = []
    for bldg_name, daily_counts in bldg_stats.items():
        total = sum(daily_counts)
        avg = round(total / len(daily_counts), 2)
        peak = max(daily_counts)
        print(f"{bldg_name}")
        print(f"  Total Clients: {total}")
        print(f"  Avg Clients/Day: {avg}")
        print(f"  Peak Clients (Single Day): {peak}")
        print("-" * 60)

        summary_data.append({
            "building": bldg_name,
            "daily_counts": daily_counts,
            "total": total,
            "average": avg,
            "peak": peak
        })

    #write_excel_report(summary_data, day_labels)

    final_report_path = write_excel_report(summary_data, day_labels)   
    # save a backup to a different directory
    backup_dir = "C:\\Users\\devon.d.youngblood\\OneDrive - US Army\\NetOps Documents\\WIRELESS REPORTS\\"
    save_report_copy(final_report_path, backup_dir)

if __name__ == "__main__":
    main()
